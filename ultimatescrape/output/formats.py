"""Consistent output rendering across every part of the system.

One row model serves swarm findings, job-board listings, GA4 reports, and graph
exports, so a CSV of GA4 geo data and a CSV of research findings have the same
shape, the same provenance columns, and the same filename convention. That
consistency is the whole point — a research team that gets a differently-shaped
spreadsheet from every command ends up rebuilding them all by hand.

Formats are chosen by the repo owner in ``uscrape.toml`` (or ``USCRAPE_OUTPUT_FORMATS``),
not hardcoded per command.
"""

from __future__ import annotations

import csv
import html as html_mod
import io
import json
import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from ..projconfig import OutputConfig, project

MEDIA_TYPES = {
    "markdown": "text/markdown",
    "json": "application/json",
    "jsonl": "application/x-ndjson",
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "html": "text/html",
    "mermaid": "text/vnd.mermaid",
}
EXTENSIONS = {**{k: f".{k}" for k in MEDIA_TYPES}, "markdown": ".md", "mermaid": ".mmd"}

#: Provenance columns. Kept last in every table so the substantive fields lead,
#: and strippable as a group via output.csv.include_provenance.
PROVENANCE = ("_verdict", "_verdict_votes", "_corroborations", "_url_status", "_agent")


@dataclass
class Artifact:
    name: str
    content: bytes
    media_type: str
    format: str

    def write(self, directory: Path) -> Path:
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / self.name
        path.write_bytes(self.content)
        return path

    @property
    def size(self) -> int:
        return len(self.content)


@dataclass
class Dataset:
    """What every renderer consumes.

    ``rows`` are flat dicts. ``sections`` optionally groups them for formats that
    can express grouping (markdown headings, xlsx sheets); formats that cannot
    just concatenate.
    """

    title: str
    rows: list[dict] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    narrative: str = ""
    sections: dict[str, list[dict]] = field(default_factory=dict)
    meta: dict[str, Any] = field(default_factory=dict)
    columns: list[str] | None = None

    def effective_columns(self, *, include_provenance: bool = True) -> list[str]:
        if self.columns:
            cols = list(self.columns)
        else:
            seen: dict[str, None] = {}
            for row in self.rows[:200]:
                for key in row:
                    seen.setdefault(key, None)
            cols = list(seen)
        substantive = [c for c in cols if not c.startswith("_")]
        provenance = [c for c in cols if c in PROVENANCE]
        return substantive + (provenance if include_provenance else [])


def slugify(value: str, max_len: int = 60) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug[:max_len] or "output"


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "; ".join(_stringify(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


# ── renderers ─────────────────────────────────────────────────────────────────


def to_markdown(data: Dataset, cfg: OutputConfig) -> str:
    out: list[str] = [f"# {data.title}\n"]
    if data.meta:
        bits = [f"{k}: {v}" for k, v in data.meta.items() if v not in (None, "")]
        if bits:
            out.append("_" + " · ".join(str(b) for b in bits) + "_\n")

    if data.summary:
        out.append("\n## Summary\n")
        out.append("| Metric | Value |")
        out.append("|---|---|")
        for key, value in data.summary.items():
            out.append(f"| {key.replace('_', ' ').title()} | {_stringify(value)} |")
        out.append("")

    if data.narrative:
        out.append("\n## Synthesis\n")
        out.append(data.narrative.strip() + "\n")

    columns = data.effective_columns()
    groups = data.sections or {"Findings": data.rows}
    for name, rows in groups.items():
        if not rows:
            continue
        out.append(f"\n## {name} ({len(rows)})\n")
        out.append("| " + " | ".join(c.lstrip("_").replace("_", " ").title() for c in columns) + " |")
        out.append("|" + "|".join("---" for _ in columns) + "|")
        for row in rows[: cfg.max_rows]:
            cells = [_stringify(row.get(c)).replace("|", "\\|").replace("\n", " ") for c in columns]
            out.append("| " + " | ".join(c[:80] for c in cells) + " |")
        if len(rows) > cfg.max_rows:
            out.append(f"\n_{len(rows) - cfg.max_rows} further rows omitted; see the JSON or CSV export._")
        out.append("")
    return "\n".join(out)


def to_json(data: Dataset, _cfg: OutputConfig) -> str:
    payload = {
        "title": data.title,
        "generated_at": datetime.now(UTC).isoformat(),
        "meta": data.meta,
        "summary": data.summary,
        "narrative": data.narrative,
        "rows": data.rows,
    }
    if data.sections:
        payload["sections"] = data.sections
    return json.dumps(payload, indent=2, ensure_ascii=False, default=str)


def to_jsonl(data: Dataset, _cfg: OutputConfig) -> str:
    return "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in data.rows)


def to_csv(data: Dataset, cfg: OutputConfig) -> str:
    columns = data.effective_columns(include_provenance=cfg.include_provenance)
    buf = io.StringIO()
    # utf-8-sig is written at encode time; Excel on Windows shows mojibake for
    # accented characters without the BOM, and the research team uses Excel.
    writer = csv.writer(buf, delimiter=cfg.delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow([c.lstrip("_") for c in columns])
    for row in data.rows:
        writer.writerow([_stringify(row.get(c)) for c in columns])
    return buf.getvalue()


def to_html(data: Dataset, cfg: OutputConfig) -> str:
    columns = data.effective_columns()
    esc = html_mod.escape

    def table(rows: Sequence[dict]) -> str:
        head = "".join(f"<th>{esc(c.lstrip('_').replace('_', ' ').title())}</th>" for c in columns)
        body = "".join(
            "<tr>"
            + "".join(
                f'<td data-col="{esc(c)}">{_linkify(_stringify(r.get(c)))}</td>' for c in columns
            )
            + "</tr>"
            for r in rows[: cfg.max_rows]
        )
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    def _linkify(value: str) -> str:
        if value.startswith("http"):
            return f'<a href="{esc(value)}" rel="noopener">{esc(value[:70])}</a>'
        return esc(value[:300])

    groups = data.sections or {"Findings": data.rows}
    sections = "".join(
        f"<h2>{esc(name)} <span class=count>{len(rows)}</span></h2>{table(rows)}"
        for name, rows in groups.items()
        if rows
    )
    summary = "".join(
        f"<div class=stat><span class=k>{esc(k.replace('_',' '))}</span>"
        f"<span class=v>{esc(_stringify(v))}</span></div>"
        for k, v in data.summary.items()
    )
    narrative = (
        f"<h2>Synthesis</h2><div class=prose>{esc(data.narrative)}</div>"
        if data.narrative
        else ""
    )
    dark = "" if cfg.theme == "light" else _DARK_CSS
    return _HTML_SHELL.format(
        title=esc(data.title),
        summary=summary,
        narrative=narrative.replace("\n", "<br>"),
        sections=sections,
        dark=dark,
        generated=datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"),
    )


def to_mermaid(data: Dataset, _cfg: OutputConfig) -> str:
    """Entity/relation diagram. Only meaningful for graph exports; for a plain
    row set it renders the strongest corroborated findings as a hub diagram."""
    lines = ["graph LR"]
    if data.meta.get("kind") == "graph":
        for row in data.rows[:200]:
            src, rel, dst = row.get("source"), row.get("relation"), row.get("target")
            if src and dst:
                lines.append(
                    f'  {_mid(src)}["{_esc_mermaid(src)}"] -->|{_esc_mermaid(rel or "related")}| '
                    f'{_mid(dst)}["{_esc_mermaid(dst)}"]'
                )
    else:
        root = _mid(data.title)
        lines.append(f'  {root}(["{_esc_mermaid(data.title)}"])')
        for row in data.rows[:40]:
            label = row.get("name") or row.get("title") or row.get("url") or "?"
            lines.append(f'  {root} --> {_mid(str(label))}["{_esc_mermaid(str(label))}"]')
    return "\n".join(lines)


def _mid(value: str) -> str:
    return "n" + re.sub(r"[^a-zA-Z0-9]", "", str(value))[:24].lower() + str(abs(hash(value)) % 9973)


def _esc_mermaid(value: str) -> str:
    return re.sub(r'["\[\]{}|<>]', " ", str(value))[:60]


def to_xlsx(data: Dataset, cfg: OutputConfig) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # noqa: F841
        raise RuntimeError(
            'xlsx output needs openpyxl — uv pip install -e ".[export]"'
        ) from None

    columns = data.effective_columns()
    wb = Workbook()
    wb.remove(wb.active)

    groups = (
        data.sections
        if (cfg.split_by_verdict and data.sections)
        else {"Data": data.rows}
    )
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for name, rows in groups.items():
        ws = wb.create_sheet(re.sub(r"[\\/*?:\[\]]", "-", name)[:31] or "Data")
        ws.append([c.lstrip("_").replace("_", " ").title() for c in columns])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            ws.append([_stringify(row.get(c))[:32000] for c in columns])

        for idx, col in enumerate(columns, start=1):
            longest = max(
                [len(col)] + [len(_stringify(r.get(col))[:60]) for r in rows[:200]] or [10]
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 12), 60)
        if cfg.freeze_header:
            ws.freeze_panes = "A2"
        if cfg.autofilter and rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    if data.summary or data.narrative:
        ws = wb.create_sheet("Summary", 0)
        ws.append(["Metric", "Value"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for key, value in data.summary.items():
            ws.append([key.replace("_", " ").title(), _stringify(value)])
        if data.narrative:
            ws.append([])
            ws.append(["Synthesis", ""])
            for line in data.narrative.splitlines():
                ws.append(["", line[:32000]])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 110
        for row in ws.iter_rows(min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_RENDERERS = {
    "markdown": to_markdown,
    "json": to_json,
    "jsonl": to_jsonl,
    "csv": to_csv,
    "html": to_html,
    "mermaid": to_mermaid,
    "xlsx": to_xlsx,
}


def render(
    data: Dataset,
    formats: Iterable[str] | None = None,
    *,
    cfg: OutputConfig | None = None,
    basename: str | None = None,
) -> list[Artifact]:
    """Render a dataset into every configured format.

    A format that fails (missing optional dependency, most often) is reported and
    skipped rather than aborting the others — losing the xlsx export should not
    cost you the markdown report too.
    """
    cfg = cfg or project.output
    chosen = list(formats) if formats else cfg.formats
    stem = basename or slugify(data.title)
    artifacts: list[Artifact] = []
    errors: list[str] = []

    for name in chosen:
        renderer = _RENDERERS.get(name)
        if renderer is None:
            errors.append(f"{name}: unknown format")
            continue
        try:
            payload = renderer(data, cfg)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{name}: {exc}")
            continue
        if isinstance(payload, str):
            # utf-8-sig for CSV so Excel on Windows reads accents correctly.
            encoding = "utf-8-sig" if name == "csv" else "utf-8"
            payload = payload.encode(encoding)
        artifacts.append(
            Artifact(
                name=f"{stem}{EXTENSIONS[name]}",
                content=payload,
                media_type=MEDIA_TYPES[name],
                format=name,
            )
        )

    if errors:
        data.meta.setdefault("render_errors", errors)
    return artifacts


def write_index(artifacts: Sequence[Artifact], data: Dataset, directory: Path) -> Path:
    lines = [f"# {data.title}\n", "Artifacts produced by this run:\n"]
    for art in sorted(artifacts, key=lambda a: a.format):
        lines.append(f"- [`{art.name}`]({art.name}) — {art.format}, {art.size:,} bytes")
    if errs := data.meta.get("render_errors"):
        lines.append("\n## Formats that failed\n")
        lines += [f"- {e}" for e in errs]
    path = directory / "index.md"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


_DARK_CSS = """
@media (prefers-color-scheme: dark){
  :root{--bg:#0b0f17;--fg:#e6edf3;--muted:#9aa7b8;--line:#1f2937;--accent:#7cc4ff;--head:#111827}
}
"""

_HTML_SHELL = """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
:root{{--bg:#fff;--fg:#111827;--muted:#6b7280;--line:#e5e7eb;--accent:#0969da;--head:#f6f8fa}}
{dark}
*{{box-sizing:border-box}}
body{{margin:0;padding:2rem 1.25rem;background:var(--bg);color:var(--fg);
 font:15px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;max-width:1200px;margin-inline:auto}}
h1{{font-size:1.7rem;margin:0 0 .25rem}}
h2{{font-size:1.15rem;margin:2rem 0 .6rem;padding-bottom:.3rem;border-bottom:1px solid var(--line)}}
.count{{color:var(--muted);font-weight:400;font-size:.85rem}}
.stats{{display:flex;flex-wrap:wrap;gap:.5rem;margin:1rem 0}}
.stat{{border:1px solid var(--line);border-radius:8px;padding:.5rem .75rem;min-width:120px}}
.stat .k{{display:block;color:var(--muted);font-size:.72rem;text-transform:uppercase;letter-spacing:.04em}}
.stat .v{{display:block;font-size:1.15rem;font-weight:600}}
.tablewrap,table{{width:100%}}
div.tablewrap{{overflow-x:auto}}
table{{border-collapse:collapse;font-size:13.5px;display:block;overflow-x:auto;white-space:nowrap}}
th,td{{border:1px solid var(--line);padding:.4rem .6rem;text-align:left;vertical-align:top}}
th{{background:var(--head);position:sticky;top:0}}
td{{max-width:420px;overflow:hidden;text-overflow:ellipsis}}
a{{color:var(--accent)}}
.prose{{white-space:pre-wrap;max-width:75ch}}
footer{{margin-top:3rem;color:var(--muted);font-size:.8rem;border-top:1px solid var(--line);padding-top:1rem}}
</style></head><body>
<h1>{title}</h1>
<div class=stats>{summary}</div>
{narrative}
{sections}
<footer>Generated by UltimateScrape · {generated}</footer>
</body></html>
"""
